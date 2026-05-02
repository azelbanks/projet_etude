"""
Notebook 22 — Evaluation sur Gold Test Set (200 posts Bluesky annotes manuellement)
====================================================================================

Objectif :
    Evaluer la performance reelle du pipeline V5 sur des posts Bluesky
    annotes independamment par 2 annotateurs (kappa = 0.808).

Resultat cle :
    Le modele detecte le SUJET (vaccin, climat, breaking news) et non
    la DESINFORMATION. Sur 200 posts reels, F1 suspect = 0.087.
    94.3% des posts etiquetes "suspect" par mots-cles sont en realite fiables.

Auteur : Azelie Bernard
Date   : Avril 2026
"""

import sys
import os
import numpy as np
import pandas as pd
import openpyxl
from sklearn.metrics import (
    classification_report, confusion_matrix, f1_score,
    accuracy_score, precision_score, recall_score
)

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))
from pipeline.expert_detector import ExpertFakeNewsDetector, EmotionFeatureExtractor

# ================================================================
#  1. CHARGEMENT DU GOLD TEST SET
# ================================================================

GOLD_PATH = os.path.join(os.path.dirname(__file__), '..', 'data',
                         'gold_test_set_annotation_completed.xlsx')
MODEL_DIR = os.path.join(os.path.dirname(__file__), '..', 'models')

print("=" * 70)
print("EVALUATION PIPELINE V5 SUR GOLD TEST SET")
print("=" * 70)

wb = openpyxl.load_workbook(GOLD_PATH)

# --- Metriques d'accord inter-annotateur ---
print("\n1. ACCORD INTER-ANNOTATEUR")
print("-" * 40)
ws_a1 = wb['Annotateur_1']
ws_a2 = wb['Annotateur_2']

labels_a1, labels_a2 = [], []
for row in ws_a1.iter_rows(min_row=2, values_only=True):
    if row[0] is not None:
        labels_a1.append(row[5])  # Label A1
for row in ws_a2.iter_rows(min_row=2, values_only=True):
    if row[0] is not None:
        labels_a2.append(row[5])  # Label A2

accords = sum(1 for a, b in zip(labels_a1, labels_a2) if a == b)
desaccords = len(labels_a1) - accords
po = accords / len(labels_a1)

# Cohen's kappa
p_fiable = (labels_a1.count('fiable') / len(labels_a1) + labels_a2.count('fiable') / len(labels_a2)) / 2
p_suspect = 1 - p_fiable
pe = p_fiable**2 + p_suspect**2
kappa = (po - pe) / (1 - pe) if pe < 1 else 0

# Gwet's AC1 (robust to prevalence paradox)
pi_f = (labels_a1.count('fiable') + labels_a2.count('fiable')) / (2 * len(labels_a1))
pe_gwet = 2 * pi_f * (1 - pi_f)
ac1 = (po - pe_gwet) / (1 - pe_gwet) if pe_gwet < 1 else 0

print(f"  Posts annotes      : {len(labels_a1)}")
print(f"  Accords            : {accords}")
print(f"  Desaccords         : {desaccords}")
print(f"  Accord brut (Po)   : {po:.3f}")
print(f"  Cohen's kappa      : {kappa:.3f} ({'substantiel' if kappa > 0.6 else 'modere'})")
print(f"  Gwet's AC1         : {ac1:.3f} (corrige le paradoxe de prevalence)")
print(f"  Annotateur 1       : {labels_a1.count('fiable')} fiable, {labels_a1.count('suspect')} suspect")
print(f"  Annotateur 2       : {labels_a2.count('fiable')} fiable, {labels_a2.count('suspect')} suspect")

# --- Chargement des labels consolides ---
ws_res = wb['Resolution']
gold_data = []
for row in ws_res.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    gold_data.append({
        'id': row[0],
        'text': row[1],
        'langue': row[2],
        'label_a1': row[3],
        'conf_a1': row[4],
        'label_a2': row[6],
        'conf_a2': row[7],
        'desaccord': row[9],
        'label_gold': row[10],
    })

df_gold = pd.DataFrame(gold_data)
print(f"\n  Label final        : {df_gold['label_gold'].value_counts().to_dict()}")
print(f"  Langues            : {df_gold['langue'].value_counts().to_dict()}")

# ================================================================
#  2. PREDICTIONS DU PIPELINE V5
# ================================================================

print("\n2. CHARGEMENT DU MODELE")
print("-" * 40)

detector = ExpertFakeNewsDetector(model_dir=MODEL_DIR, threshold=0.44)
detector.load(suffix='expert_v5')
emo = EmotionFeatureExtractor(model_dir=MODEL_DIR)
emo.load()
print(f"  Modele  : Pipeline V5 (TF-IDF + 15 features linguistiques + 7 emotions)")
print(f"  Seuil   : 0.44")

print("\n  Inference en cours...")
texts = pd.Series(df_gold['text'].values)
results = detector.predict(texts)

y_true = (df_gold['label_gold'] == 'suspect').astype(int).values
y_pred = results['prediction_label'].values
scores = results['ai_score_credibility'].values

# ================================================================
#  3. METRIQUES GLOBALES
# ================================================================

print("\n3. METRIQUES GLOBALES")
print("-" * 40)

acc = accuracy_score(y_true, y_pred)
f1_macro = f1_score(y_true, y_pred, average='macro')
f1_suspect = f1_score(y_true, y_pred, pos_label=1, zero_division=0)
f1_fiable = f1_score(y_true, y_pred, pos_label=0)
prec_suspect = precision_score(y_true, y_pred, pos_label=1, zero_division=0)
rec_suspect = recall_score(y_true, y_pred, pos_label=1, zero_division=0)

print(f"  Accuracy       : {acc:.3f}")
print(f"  F1 macro       : {f1_macro:.3f}")
print(f"  F1 fiable      : {f1_fiable:.3f}")
print(f"  F1 suspect     : {f1_suspect:.3f}")
print(f"  Precision susp.: {prec_suspect:.3f}")
print(f"  Recall suspect : {rec_suspect:.3f}")

print(f"\n  Classification Report:")
print(classification_report(y_true, y_pred, target_names=['fiable', 'suspect']))

cm = confusion_matrix(y_true, y_pred)
print(f"  Matrice de confusion:")
print(f"                    Pred fiable  Pred suspect")
print(f"    Gold fiable  :  {cm[0][0]:>10}  {cm[0][1]:>12}")
print(f"    Gold suspect :  {cm[1][0]:>10}  {cm[1][1]:>12}")

# ================================================================
#  4. ANALYSE PAR LANGUE
# ================================================================

print("\n4. ANALYSE PAR LANGUE")
print("-" * 40)

for lang in ['fr', 'en']:
    mask = df_gold['langue'].values == lang
    yt, yp = y_true[mask], y_pred[mask]
    n = mask.sum()
    n_gold_suspect = yt.sum()
    n_pred_suspect = yp.sum()
    a = accuracy_score(yt, yp)
    f1 = f1_score(yt, yp, average='macro', zero_division=0)

    print(f"\n  {lang.upper()} ({n} posts, {n_gold_suspect} suspects gold, {n_pred_suspect} suspects pred)")
    print(f"    Accuracy: {a:.3f}, F1 macro: {f1:.3f}")

    fp = ((yt == 0) & (yp == 1)).sum()
    fn = ((yt == 1) & (yp == 0)).sum()
    print(f"    Faux positifs (fiable->suspect): {fp}")
    print(f"    Faux negatifs (suspect->fiable): {fn}")

# ================================================================
#  5. ANALYSE DES ERREURS
# ================================================================

print("\n5. ANALYSE DES ERREURS")
print("-" * 40)

fp_mask = (y_true == 0) & (y_pred == 1)
fn_mask = (y_true == 1) & (y_pred == 0)

print(f"\n  Faux positifs (fiable classe suspect): {fp_mask.sum()}/191 ({fp_mask.sum()/191*100:.1f}%)")
print(f"  --- Exemples ---")
for idx in np.where(fp_mask)[0][:5]:
    text = df_gold.iloc[idx]['text'][:120].replace('\n', ' ')
    score = scores[idx]
    lang = df_gold.iloc[idx]['langue']
    print(f"    [{lang}] score={score:.3f} | {text}...")

print(f"\n  Faux negatifs (suspect classe fiable): {fn_mask.sum()}/9 ({fn_mask.sum()/9*100:.1f}%)")
print(f"  --- Exemples ---")
for idx in np.where(fn_mask)[0]:
    text = df_gold.iloc[idx]['text'][:120].replace('\n', ' ')
    score = scores[idx]
    lang = df_gold.iloc[idx]['langue']
    print(f"    [{lang}] score={score:.3f} | {text}...")

# ================================================================
#  6. DISTRIBUTION DES SCORES
# ================================================================

print("\n6. DISTRIBUTION DES SCORES")
print("-" * 40)

scores_fiable = scores[y_true == 0]
scores_suspect = scores[y_true == 1]

print(f"  Gold fiable  : moyenne={scores_fiable.mean():.3f}, std={scores_fiable.std():.3f}, "
      f"median={np.median(scores_fiable):.3f}")
print(f"  Gold suspect : moyenne={scores_suspect.mean():.3f}, std={scores_suspect.std():.3f}, "
      f"median={np.median(scores_suspect):.3f}")
print(f"\n  Ecart des moyennes : {abs(scores_fiable.mean() - scores_suspect.mean()):.3f}")
print(f"  => Le modele ne distingue quasiment PAS fiable vs suspect sur Bluesky")

print(f"\n  Predictions du modele:")
print(f"    Classes fiable : {(y_pred==0).sum()}/200 ({(y_pred==0).sum()/2:.1f}%)")
print(f"    Classes suspect: {(y_pred==1).sum()}/200 ({(y_pred==1).sum()/2:.1f}%)")

# ================================================================
#  7. COMPARAISON METRIQUES SYNTHETIQUES vs GOLD
# ================================================================

print("\n7. COMPARAISON SYNTHETIQUE vs GOLD")
print("-" * 40)

print("""
  +----------------------+----------------+----------------+
  | Metrique             | Synthetique    | Gold (reel)    |
  +----------------------+----------------+----------------+
  | F1 global (macro)    | 0.913          | {f1_macro:.3f}          |
  | F1 suspect           | ~0.90          | {f1_s:.3f}          |
  | Accuracy             | 0.93           | {acc:.3f}          |
  | % fiable             | 73.4%          | {pct:.1f}%          |
  +----------------------+----------------+----------------+

  CONCLUSION : Les metriques sur donnees synthetiques (F1=0.913)
  sont mecaniquement gonflees par un biais thematique dans les
  datasets d'entrainement. Sur des posts Bluesky reels annotes,
  le F1 suspect chute a {f1_s:.3f}.

  Le modele apprend a detecter le SUJET (vaccin, climat, complot)
  et non la DESINFORMATION. Un post fiable sur le climat est
  classe suspect parce qu'il contient des mots-cles correles
  aux fake news dans le dataset d'entrainement.
""".format(f1_macro=f1_macro, f1_s=f1_suspect, acc=acc,
           pct=(y_pred==0).sum()/2))

# ================================================================
#  8. PISTES D'AMELIORATION
# ================================================================

print("8. PISTES D'AMELIORATION")
print("-" * 40)
print("""
  1. DONNEES : Utiliser le gold test set (200 posts) comme seed pour
     creer un dataset d'entrainement specifique Bluesky annote.

  2. FEATURES : Ajouter des features basees sur le registre enonciatif
     (presence de sources, marqueurs d'opinion, ton interrogatif vs
     affirmatif) plutot que le champ lexical.

  3. MODELE : Fine-tuner CamemBERT/RoBERTa directement sur des posts
     Bluesky annotes (meme un petit dataset de 500-1000 posts).

  4. SEUIL : Ajuster le seuil de decision sur le gold set plutot que
     sur les donnees synthetiques.

  5. ENSEMBLE : Combiner le score TF-IDF avec un score de presence
     de source (URL vers media reconnu, citation d'expert).
""")

print("=" * 70)
print("FIN DE L'EVALUATION GOLD TEST SET")
print("=" * 70)
