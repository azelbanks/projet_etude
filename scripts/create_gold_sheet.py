"""
Genere un fichier Excel (.xlsx) pret pour l'annotation du Gold Test Set.

- Menus deroulants pour label (fiable/suspect) et confiance (1/2/3)
- Mise en forme conditionnelle (vert=fiable, rouge=suspect)
- Colonnes verrouillees pour le texte, editables pour l'annotation
- 2 onglets : Annotateur_1 et Annotateur_2

Si MongoDB est accessible : extrait 200 posts stratifies.
Sinon : cree le template vide a remplir apres extraction CSV.

Usage:
    python scripts/create_gold_sheet.py [--from-csv data/gold_test_set_blind.csv]
    python scripts/create_gold_sheet.py [--from-mongo --n_posts 200]

Auteur: Azelie Bernard - Thumalien Team
"""

import argparse
import sys
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


def create_annotation_sheet(ws, df, annotateur_num):
    """
    Configure un onglet d'annotation avec menus deroulants.
    """
    # --- Couleurs ---
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    TEXT_FONT = Font(name="Arial", size=10)
    ANNOT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # jaune clair
    GREEN_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # --- En-tetes ---
    headers = [
        ("ID", 6),
        ("Texte du post", 80),
        ("Langue", 8),
        ("Mots", 7),
        ("Strate", 18),
        (f"Label (A{annotateur_num})", 15),
        (f"Confiance (A{annotateur_num})", 15),
        (f"Justification (A{annotateur_num})", 35),
    ]

    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Hauteur header
    ws.row_dimensions[1].height = 30

    # --- Donnees ---
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        # ID
        ws.cell(row=row_idx, column=1, value=row.get("id", row_idx - 1)).font = TEXT_FONT
        # Texte
        cell_text = ws.cell(row=row_idx, column=2, value=str(row.get("text", "")))
        cell_text.font = TEXT_FONT
        cell_text.alignment = Alignment(wrap_text=True, vertical="top")
        # Langue
        ws.cell(row=row_idx, column=3, value=row.get("langue", "")).font = TEXT_FONT
        # Mots
        ws.cell(row=row_idx, column=4, value=row.get("nb_mots", 0)).font = TEXT_FONT
        # Strate
        ws.cell(row=row_idx, column=5, value=row.get("strate", "")).font = TEXT_FONT

        # Colonnes annotation (fond jaune)
        for col in [6, 7, 8]:
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = ANNOT_FILL
            cell.font = TEXT_FONT

        # Bordures
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = BORDER

        # Hauteur adaptee
        n_mots = row.get("nb_mots", 10)
        ws.row_dimensions[row_idx].height = max(25, min(80, n_mots * 2))

    n_rows = len(df) + 1  # +1 pour header

    # --- Menu deroulant : Label (fiable / suspect) ---
    dv_label = DataValidation(
        type="list",
        formula1='"fiable,suspect"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Choisir: fiable ou suspect",
        showInputMessage=True,
        promptTitle="Label",
        prompt="fiable = post credible/normal\nsuspect = desinformation/manipulation"
    )
    dv_label.add(f"F2:F{n_rows}")
    ws.add_data_validation(dv_label)

    # --- Menu deroulant : Confiance (1, 2, 3) ---
    dv_conf = DataValidation(
        type="list",
        formula1='"1,2,3"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error="Choisir: 1 (pas sur), 2 (assez sur), 3 (certain)",
        showInputMessage=True,
        promptTitle="Confiance",
        prompt="1 = pas sur\n2 = assez sur\n3 = certain"
    )
    dv_conf.add(f"G2:G{n_rows}")
    ws.add_data_validation(dv_conf)

    # --- Mise en forme conditionnelle ---
    ws.conditional_formatting.add(
        f"F2:F{n_rows}",
        CellIsRule(operator="equal", formula=['"fiable"'], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        f"F2:F{n_rows}",
        CellIsRule(operator="equal", formula=['"suspect"'], fill=RED_FILL)
    )

    # --- Figer la premiere ligne + colonne texte ---
    ws.freeze_panes = "C2"

    # --- Onglet Guidelines ---
    return ws


def create_guidelines_sheet(wb):
    """Cree un onglet avec les consignes d'annotation."""
    ws = wb.create_sheet("Guidelines", 0)

    TITLE_FONT = Font(name="Arial", size=14, bold=True, color="2C3E50")
    SECTION_FONT = Font(name="Arial", size=12, bold=True, color="2C3E50")
    TEXT_FONT = Font(name="Arial", size=11)
    BOLD_FONT = Font(name="Arial", size=11, bold=True)
    GREEN = Font(name="Arial", size=11, color="2E7D32")
    RED = Font(name="Arial", size=11, color="C62828")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 90

    row = 2
    ws.cell(row=row, column=2, value="GOLD TEST SET - Consignes d'annotation").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=2, value="Projet Thumalien - Detection de desinformation sur Bluesky").font = TEXT_FONT
    row += 2

    ws.cell(row=row, column=2, value="OBJECTIF").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=2,
            value="Annoter chaque post comme 'fiable' ou 'suspect' SANS connaitre la prediction de l'IA.").font = TEXT_FONT
    row += 1
    ws.cell(row=row, column=2,
            value="Ce gold set servira a evaluer la qualite reelle des modeles sur des donnees non synthetiques.").font = TEXT_FONT
    row += 2

    ws.cell(row=row, column=2, value="DEFINITIONS").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=2,
            value="SUSPECT = desinformation, theorie du complot, manipulation, affirmation fausse non sourcee").font = RED
    row += 1
    ws.cell(row=row, column=2,
            value="FIABLE = post factuel, source, opinion/expression normale, humour/satire identifiable").font = GREEN
    row += 2

    ws.cell(row=row, column=2, value="REGLES DE DECISION").font = SECTION_FONT
    row += 1
    rules = [
        "1. Lire le post en entier avant de decider",
        "2. Se demander : 'Ce post propage-t-il une information fausse ou manipulatrice ?'",
        "3. Un post qui POSE UNE QUESTION sans affirmer → fiable",
        "4. Un post SATIRIQUE clairement identifiable → fiable",
        "5. Un post qui CITE UNE SOURCE FIABLE meme sur sujet polemique → fiable",
        "6. Un post qui AFFIRME SANS SOURCE un fait verifiable → suspect",
        "7. Vocabulaire alarmiste/complotiste (on nous cache, reveillez-vous, ils ne veulent pas...) → suspect",
        "8. En cas de doute sincere → annoter quand meme, mettre confiance=1",
    ]
    for rule in rules:
        ws.cell(row=row, column=2, value=rule).font = TEXT_FONT
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="CONFIANCE").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=2, value="1 = pas sur (je devine, le post est ambigu)").font = TEXT_FONT
    row += 1
    ws.cell(row=row, column=2, value="2 = assez sur (je pense avoir raison mais c'est discutable)").font = TEXT_FONT
    row += 1
    ws.cell(row=row, column=2, value="3 = certain (aucun doute, le post est clairement fiable ou suspect)").font = TEXT_FONT
    row += 2

    ws.cell(row=row, column=2, value="PROCEDURE").font = SECTION_FONT
    row += 1
    steps = [
        "1. Commencer par l'onglet 'Annotateur_1' (ou 'Annotateur_2' si vous etes le 2e)",
        "2. Pour chaque ligne : lire le texte → choisir label → choisir confiance → justification optionnelle",
        "3. Temps estime : ~1h pour 200 posts (3-4 sec/post)",
        "4. Faire des pauses toutes les 50 posts pour eviter la fatigue",
        "5. NE PAS revenir en arriere modifier systematiquement (biais de coherence)",
    ]
    for step in steps:
        ws.cell(row=row, column=2, value=step).font = TEXT_FONT
        row += 1

    ws.freeze_panes = "A1"
    return ws


def create_resolution_sheet(wb, df):
    """Cree un onglet pour la resolution des desaccords."""
    ws = wb.create_sheet("Resolution")

    HEADER_FILL = PatternFill(start_color="5C6BC0", end_color="5C6BC0", fill_type="solid")
    HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    TEXT_FONT = Font(name="Arial", size=10)
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        ("ID", 6), ("Texte", 80), ("Langue", 8),
        ("Label A1", 12), ("Conf. A1", 10), ("Justif. A1", 25),
        ("Label A2", 12), ("Conf. A2", 10), ("Justif. A2", 25),
        ("Desaccord?", 12), ("Label final", 15), ("Commentaire", 35),
    ]

    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Donnees (juste ID et texte, le reste sera copie apres annotation)
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row.get("id", row_idx - 1)).font = TEXT_FONT
        cell_text = ws.cell(row=row_idx, column=2, value=str(row.get("text", "")))
        cell_text.font = TEXT_FONT
        cell_text.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=3, value=row.get("langue", "")).font = TEXT_FONT
        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).border = BORDER

    n_rows = len(df) + 1

    # Menu deroulant label final
    dv = DataValidation(type="list", formula1='"fiable,suspect"', allow_blank=True)
    dv.add(f"K2:K{n_rows}")
    ws.add_data_validation(dv)

    # Formule desaccord (=SI(D2<>G2, "OUI", ""))
    for row_idx in range(2, n_rows + 1):
        ws.cell(row=row_idx, column=10).value = f'=IF(AND(D{row_idx}<>"",G{row_idx}<>""),IF(D{row_idx}<>G{row_idx},"OUI",""),"")'

    ws.freeze_panes = "C2"
    return ws


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--from-csv", type=str, default=None,
                        help="Charger depuis un CSV existant")
    parser.add_argument("--from-mongo", action="store_true",
                        help="Extraire depuis MongoDB")
    parser.add_argument("--n_posts", type=int, default=200)
    parser.add_argument("--output", type=str,
                        default="data/gold_test_set_annotation.xlsx")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    # --- Charger les donnees ---
    if args.from_csv:
        print(f"Chargement depuis {args.from_csv}...")
        df = pd.read_csv(args.from_csv)
    elif args.from_mongo:
        print("Extraction depuis MongoDB...")
        import random
        random.seed(args.seed)
        sys.path.insert(0, os.path.dirname(__file__))
        from extract_gold_test_set import get_mongo_collection, extract_stratified_sample
        collection = get_mongo_collection()
        posts = extract_stratified_sample(collection, args.n_posts)
        # Convertir en DataFrame
        rows = []
        for i, p in enumerate(posts, 1):
            rows.append({
                "id": i,
                "text": p.get("text", ""),
                "langue": p.get("ai_language", ""),
                "nb_mots": p.get("text_word_count", 0),
                "strate": p.get("strate", ""),
            })
        df = pd.DataFrame(rows)
        # Sauvegarder aussi en CSV
        csv_path = args.output.replace(".xlsx", ".csv")
        df.to_csv(csv_path, index=False)
        print(f"CSV sauvegarde: {csv_path}")
    else:
        print("Erreur: specifier --from-csv ou --from-mongo")
        sys.exit(1)

    if "id" not in df.columns:
        df.insert(0, "id", range(1, len(df) + 1))

    print(f"Posts charges: {len(df)}")

    # --- Creer le workbook ---
    wb = Workbook()
    # Supprimer la feuille par defaut
    wb.remove(wb.active)

    # 1. Guidelines
    create_guidelines_sheet(wb)

    # 2. Annotateur 1
    ws1 = wb.create_sheet("Annotateur_1")
    create_annotation_sheet(ws1, df, annotateur_num=1)
    ws1.sheet_properties.tabColor = "4CAF50"

    # 3. Annotateur 2
    ws2 = wb.create_sheet("Annotateur_2")
    create_annotation_sheet(ws2, df, annotateur_num=2)
    ws2.sheet_properties.tabColor = "2196F3"

    # 4. Resolution
    create_resolution_sheet(wb, df)

    # --- Sauvegarder ---
    wb.save(args.output)
    print(f"\nFichier cree: {args.output}")
    print(f"\nOnglets:")
    print(f"  1. Guidelines      - Consignes d'annotation (lire en premier)")
    print(f"  2. Annotateur_1    - Onglet vert (toi)")
    print(f"  3. Annotateur_2    - Onglet bleu (2e annotateur)")
    print(f"  4. Resolution      - Pour resoudre les desaccords apres")
    print(f"\nPour importer dans Google Sheets:")
    print(f"  Google Drive → Nouveau → Importer → {args.output}")
    print(f"  Les menus deroulants seront preserves.")


if __name__ == "__main__":
    main()
